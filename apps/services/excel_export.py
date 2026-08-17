"""Xuất list[dict] thành file .xlsx với sheet tuỳ chỉnh."""

from __future__ import annotations

import re
from typing import Iterable, List


def sanitize_sheet_name(name: str, default: str = "Sheet1") -> str:
    cleaned = re.sub(r"[^A-Za-z0-9 _-]+", "_", name).strip() or default
    return cleaned[:31]


# Màu NỀN cảnh báo cho các ô cần chú ý (vd. sheet Summary). PHẢI là aRGB 8 ký
# tự: openpyxl đệm chuỗi 6 ký tự bằng "00" ở ĐẦU, tức "FF0000" ra
# rgb="00FF0000" (alpha 0 — vô hình), không phải màu đỏ mong muốn.
_ALERT_FILL_COLOR = "FFFF0000"

_alert_fill_cache = None


def _alert_fill():
    """Nền đỏ dùng chung cho mọi ô cảnh báo. Tạo đúng MỘT lần.

    Tô NỀN chứ không phải màu chữ: nền vẫn hiện được trên ô KHÔNG có chữ (vd.
    ngày trống của một nghiên cứu 0 bản ghi), trong khi chữ đỏ trên ô trống thì
    không có gì để tô — không có tác dụng cảnh báo.

    openpyxl gom các PatternFill giống nhau vào chung một entry trong bảng
    style của workbook, nên hàng nghìn ô cùng trỏ về một Fill instance chỉ tốn
    một dòng trong ``styles.xml``; cấp Fill mới cho từng ô vừa phí vừa làm
    phình bảng style không cần thiết. Import openpyxl vẫn nằm trong hàm, giống
    mọi chỗ khác trong repo — openpyxl chỉ nạp khi thật sự xuất Excel.
    """

    global _alert_fill_cache
    if _alert_fill_cache is None:
        from openpyxl.styles import PatternFill
        _alert_fill_cache = PatternFill(
            fill_type="solid", start_color=_ALERT_FILL_COLOR, end_color=_ALERT_FILL_COLOR,
        )
    return _alert_fill_cache


def append_row_as_text(ws, values: Iterable, alerts: Iterable[bool] | None = None) -> None:
    """Ghi 1 dòng, ép các ô bị hiểu nhầm là công thức về dạng text.

    openpyxl gán ``data_type="f"`` cho mọi chuỗi bắt đầu bằng ``=`` (và chỉ
    ``=``; ``+ - @`` đã được lưu dạng text sẵn trong .xlsx). Hệ quả: kết quả
    xét nghiệm dạng ``=<0.5`` mở ra thành ``#NAME?`` — mất dữ liệu âm thầm —
    còn chuỗi kiểu ``=cmd|'/c calc'!A1`` trở thành lệnh DDE.

    Đặt ``data_type="s"`` giữ nguyên **đúng văn bản gốc**. Không dùng cách
    prefix dấu ``'`` như bên CSV: ở .xlsx nó sẽ làm hỏng các giá trị âm hợp lệ
    (``-5.2``) vốn không hề bị coi là công thức.

    Ô được dựng sẵn TRƯỚC khi append, không sửa lại sau. Cách cũ (``ws.append``
    rồi duyệt ``ws[ws.max_row]``) là O(n²): cả ``max_row`` lẫn ``ws[<int>]``
    đều quét toàn bộ ``ws._cells`` mỗi lần gọi, nên chi phí tăng theo bình
    phương số dòng — đo được 0,37s cho 500 dòng nhưng 21s cho 4.000 dòng, tức
    hàng chục phút với một lô XML4 thật. Cách này O(số cột) mỗi dòng và chạy
    được cả ở chế độ ``write_only`` (nơi không thể truy cập lại ô đã ghi).

    ``alerts`` (tuỳ chọn) là danh sách cờ bool CÙNG ĐỘ DÀI với ``values`` — ô
    nào có cờ True thì tô NỀN đỏ. Nhận cờ đã tính sẵn thay vì một hàm predicate
    vì caller (vd. sheet Summary) cần áp quy tắc theo TỪNG CỘT (một cột được
    miễn tô đỏ bất kể giá trị) — quyết định đó phải làm trước khi giá trị tới
    được hàm này. Hai lớp bảo vệ ĐỘC LẬP và có thể cùng áp lên một ô: chuỗi vừa
    bắt đầu bằng ``=`` vừa bị đánh dấu cảnh báo thì vẫn giữ ``data_type="s"``
    VÀ được tô nền đỏ. Ở ``write_only`` không thể quay lại sửa ô đã ghi, nên
    style bắt buộc phải gán ngay lúc dựng ô — không có phương án "tô màu sau".
    """

    from openpyxl.cell import Cell

    fill = _alert_fill() if alerts is not None else None
    alerts = list(alerts) if alerts is not None else ()

    row = []
    for i, value in enumerate(values):
        is_formula = isinstance(value, str) and value.startswith("=")
        is_alert = fill is not None and i < len(alerts) and alerts[i]
        if not (is_formula or is_alert):
            row.append(value)
            continue
        cell = Cell(worksheet=ws, column=1, row=1, value=value)
        if is_formula:
            cell.data_type = "s"  # toạ độ thật do ws.append gán lại
        if is_alert:
            cell.fill = fill
        row.append(cell)
    ws.append(row)


def collect_columns(rows: Iterable[dict]) -> List[str]:
    return list(dict.fromkeys(key for row in rows for key in row))


def write_rows_to_xlsx(
    path: str,
    rows: List[dict],
    sheet_name: str = "Sheet1",
    columns: List[str] | None = None,
) -> int:
    """Ghi rows ra file Excel. Trả về số dòng dữ liệu ghi được."""

    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    ws.title = sanitize_sheet_name(sheet_name)

    cols = columns or collect_columns(rows)
    append_row_as_text(ws, cols)
    for row in rows:
        append_row_as_text(ws, [row.get(column, "") for column in cols])

    wb.save(path)
    return len(rows)


__all__ = [
    "append_row_as_text",
    "collect_columns",
    "sanitize_sheet_name",
    "write_rows_to_xlsx",
]
