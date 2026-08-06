"""File Excel danh sách nghiên cứu — gắn ``USUBJID`` và lọc theo khoảng ngày.

Đây là loại mapping thứ BA trong dự án, đừng lẫn với hai loại kia:

===================  ====================  ====================  ===================
                     medical_catalog       mapping_excel         study_mapping
===================  ====================  ====================  ===================
Nguồn                dữ liệu phát hành     người dùng chọn       người dùng chọn
Tiêu đề cột          cố định trong code    do người dùng đặt     cố định (USUBJID,
                                                                 EMR_ID, ngày)
Vai trò              lọc Include/Exclude   chỉ gắn thêm cột      gắn USUBJID, lọc
                                                                 theo ngày, ẩn định
                                                                 danh gốc
===================  ====================  ====================  ===================

Quy ước file (dòng 1 là tiêu đề, thứ tự cột không quan trọng):

- ``USUBJID`` — mã nghiên cứu, là cột DUY NHẤT được đưa vào file xuất ra.
- ``EMR_ID`` — định danh bệnh nhân, dùng để ghép với hồ sơ XML4 theo hai cách:
  khớp TOÀN BỘ với ``MA_LK``, hoặc khớp **10 ký tự cuối** với ``ID`` / ``ID_GOC``.
  Hai cách này được thử theo đúng thứ tự đó (xem ``xml_to_excel._match_study``).
  (Tên cột từng là ``STUDY_ID``/``HRN``; đổi tên để khớp chuẩn CDISC USUBJID và
  quy ước EMR nội bộ — xem yêu cầu đổi tên trong lịch sử commit.)
- Cột ngày bắt đầu (``START_DATE``) và ngày kết thúc (``END_DATE`` / ``FROM_DATE``
  / ``TO_DATE``…) là TUỲ CHỌN: có giá trị thì lọc, bỏ trống thì lấy hết.

Vì sao nhận nhiều tên cho cột ngày kết thúc: file thực tế của người dùng đặt tên
cột thứ hai là ``FROM_DATE`` nhưng ý nghĩa là ngày KẾT THÚC. Cột bắt đầu chỉ
nhận đúng ``START_DATE`` nên không có chuyện hai cột cùng tranh một vai; tên cột
thực sự dùng được ghi vào log mỗi lần nạp để người dùng đối chiếu được.

**Lọc theo ngày chỉ so phần yyyymmdd**, cố ý bỏ phần giờ của ``NGAY_KQ``: người
dùng yêu cầu tính trọn 24 giờ của cả hai đầu mút, nên một kết quả lúc 23:50 ngày
cuối vẫn phải được giữ. So ngày với ngày là cách duy nhất không phụ thuộc vào
việc ``NGAY_KQ`` có bao nhiêu chữ số giờ/phút/giây.

Ngày ghi sai trong file mapping thì DỪNG HẲN, không âm thầm bỏ qua: một khoảng
ngày bị lờ đi có nghĩa là dữ liệu ra nhiều hơn phạm vi nghiên cứu mà không ai
biết. Ngược lại, một bản ghi XML4 khớp EMR_ID nhưng NGAY_KQ nằm NGOÀI khoảng
ngày của mapping thì bị loại bỏ HẲN — không ghi vào bất kỳ sheet nào (xem
``xml_to_excel._split_by_study``), vì với dữ liệu nghiên cứu, ở ngoài phạm vi
thời gian nghĩa là không thuộc nghiên cứu, không phải "cần rà soát thêm".
"""

from __future__ import annotations

import logging
import re
from dataclasses import dataclass
from datetime import date, datetime
from pathlib import Path
from types import MappingProxyType
from typing import Dict, List, Mapping, Optional, Tuple

# Dùng chung quy ước định dạng file và cách chuẩn hoá khoá với mapping_excel:
# hai file này người dùng chọn từ cùng một nút, "1234567890 " phải khớp giống
# nhau ở cả hai chế độ.
from apps.services.mapping_excel import (
    LEGACY_SUFFIXES,
    SUPPORTED_SUFFIXES,
    normalize_key,
)

log = logging.getLogger(__name__)

# Tên cột bắt buộc. Thiếu một trong hai thì file không phải danh sách nghiên
# cứu và caller sẽ quay về mapping tổng quát (mapping_excel).
COL_STUDY_ID = "USUBJID"
COL_HRN = "EMR_ID"

# Cột ngày bắt đầu chỉ nhận đúng một tên; ngày kết thúc nhận nhiều biến thể.
START_DATE_HEADERS = ("START_DATE",)
END_DATE_HEADERS = ("END_DATE", "FROM_DATE", "TO_DATE", "THRU_DATE", "FINISH_DATE")

# Số ký tự cuối của ID / ID_GOC được đem so với HRN.
ID_SUFFIX_LENGTH = 10

# Số khoá trùng tối đa ghi ra log.
_MAX_LOGGED_DUPLICATES = 10

_DIGITS = re.compile(r"\d+")

# Các định dạng ngày dạng chữ hay gặp trong file người dùng gõ tay.
_TEXT_DATE_FORMATS = ("%d/%m/%Y", "%d-%m-%Y", "%Y-%m-%d", "%Y/%m/%d", "%d.%m.%Y")


class StudyMappingError(Exception):
    """File danh sách nghiên cứu không dùng được. Message cho người dùng đọc."""


@dataclass(frozen=True)
class StudyRow:
    """Một dòng trong danh sách nghiên cứu.

    ``start_date`` / ``end_date`` là số nguyên ``yyyymmdd`` (hoặc ``None`` nếu
    ô để trống). Giữ dạng số nguyên vì so sánh ngày lúc này chỉ là so hai số —
    rẻ hơn hẳn dựng ``datetime`` cho từng bản ghi trong lô hàng trăm nghìn dòng.
    """

    study_id: str
    start_date: Optional[int] = None
    end_date: Optional[int] = None

    @property
    def has_range(self) -> bool:
        return self.start_date is not None or self.end_date is not None

    def covers(self, day: int) -> bool:
        """``day`` (yyyymmdd) có nằm trong khoảng không. Thiếu đầu mút = mở."""

        if self.start_date is not None and day < self.start_date:
            return False
        if self.end_date is not None and day > self.end_date:
            return False
        return True


@dataclass(frozen=True)
class StudyMapping:
    """Danh sách nghiên cứu đã đọc xong.

    ``by_hrn`` tra theo HRN đầy đủ (dùng cho ``MA_LK``); ``by_suffix`` tra theo
    10 ký tự cuối của HRN (dùng cho ``ID`` / ``ID_GOC``). Hai chỉ mục tách rời
    để cách khớp "toàn bộ" không bao giờ vô tình ăn theo kết quả của cách khớp
    "10 ký tự cuối".
    """

    by_hrn: Mapping[str, StudyRow]
    by_suffix: Mapping[str, StudyRow]
    source_name: str
    start_column: Optional[str] = None
    end_column: Optional[str] = None

    @property
    def has_date_filter(self) -> bool:
        return self.start_column is not None or self.end_column is not None

    def lookup_full(self, value) -> Optional[StudyRow]:
        """Khớp TOÀN BỘ giá trị với HRN (dùng cho ``MA_LK``)."""

        key = normalize_key(value)
        return self.by_hrn.get(key) if key else None

    def lookup_suffix(self, value) -> Optional[StudyRow]:
        """Khớp 10 ký tự cuối của giá trị với HRN (dùng cho ``ID``/``ID_GOC``)."""

        key = normalize_key(value)
        if len(key) < ID_SUFFIX_LENGTH:
            # Ngắn hơn 10 ký tự thì không đủ căn cứ để khẳng định trùng — bỏ,
            # thà không khớp còn hơn gán nhầm USUBJID cho bệnh nhân khác.
            return None
        return self.by_suffix.get(key[-ID_SUFFIX_LENGTH:])

    def __len__(self) -> int:
        return len(self.by_hrn)


def parse_record_date(value) -> Optional[int]:
    """Đọc ``NGAY_KQ`` của XML4 thành số ``yyyymmdd``. ``None`` nếu không đọc được.

    ``NGAY_KQ`` là chuỗi số dạng ``yyyymmddhhmm`` (đôi khi có thêm giây, đôi khi
    chỉ có ngày). Chỉ 8 chữ số đầu được dùng — phần giờ bị bỏ có chủ đích, xem
    docstring module.
    """

    text = "".join(_DIGITS.findall(str(value or "")))
    if len(text) < 8:
        return None
    try:
        day = int(text[:8])
    except ValueError:  # không thể xảy ra sau findall, phòng xa
        return None
    return day if _is_valid_yyyymmdd(day) else None


def _is_valid_yyyymmdd(day: int) -> bool:
    try:
        datetime.strptime(str(day), "%Y%m%d")
    except ValueError:
        return False
    return True


def _cell_display(value) -> str:
    """Ép ô Excel về chuỗi hiển thị, giữ nguyên hoa/thường.

    openpyxl trả số nguyên dưới dạng float (``1054.0``) — để nguyên thì
    ``USUBJID`` số sẽ ra ``"1054.0"`` trong file xuất.
    """

    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def _cell_key(value) -> str:
    """Ép ô Excel về chuỗi khoá (đã chuẩn hoá hoa/thường, bỏ khoảng trắng)."""

    return normalize_key(_cell_display(value))


def _parse_date_cell(value, column: str, row_no: int) -> Optional[int]:
    """Đọc ô ngày trong file mapping thành ``yyyymmdd``. Ô trống trả ``None``.

    Ném ``StudyMappingError`` nếu có giá trị nhưng không hiểu được: bỏ qua âm
    thầm sẽ làm bản ghi lọt ra ngoài khoảng nghiên cứu mà không ai hay.
    """

    if value is None:
        return None
    if isinstance(value, datetime):
        return value.year * 10000 + value.month * 100 + value.day
    if isinstance(value, date):
        return value.year * 10000 + value.month * 100 + value.day

    if isinstance(value, (int, float)) and not isinstance(value, bool):
        if isinstance(value, float) and value.is_integer():
            value = int(value)
        digits = str(value)
    else:
        digits = str(value).strip()
        if not digits:
            return None
        for fmt in _TEXT_DATE_FORMATS:
            try:
                parsed = datetime.strptime(digits, fmt)
            except ValueError:
                continue
            return parsed.year * 10000 + parsed.month * 100 + parsed.day
        digits = "".join(_DIGITS.findall(digits))

    if len(digits) >= 8:
        day = int(digits[:8])
        if _is_valid_yyyymmdd(day):
            return day

    raise StudyMappingError(
        f"dòng {row_no}, cột {column}: không hiểu giá trị ngày {value!r}.\n"
        f"      Dùng định dạng ngày của Excel, hoặc dd/mm/yyyy, hoặc yyyymmdd."
    )


def _index_headers(headers: List[str]) -> Dict[str, int]:
    """Ánh xạ tên cột (đã chuẩn hoá) → chỉ số cột, giữ cột xuất hiện đầu tiên."""

    index: Dict[str, int] = {}
    for i, name in enumerate(headers):
        key = normalize_key(name)
        if key and key not in index:
            index[key] = i
    return index


def _pick_column(index: Mapping[str, int], candidates) -> Tuple[Optional[str], Optional[int]]:
    for name in candidates:
        if name in index:
            return name, index[name]
    return None, None


def _cell_at(row, position: Optional[int]):
    if position is None or position >= len(row):
        return None
    return row[position]


def load_study_mapping(path) -> Optional[StudyMapping]:
    """Nạp file danh sách nghiên cứu.

    Trả ``None`` nếu file Excel hợp lệ nhưng KHÔNG có đủ cột ``USUBJID`` và
    ``EMR_ID`` — khi đó caller quay về mapping tổng quát ``mapping_excel``. Ném
    ``StudyMappingError`` nếu file không mở/không đọc được, hoặc có cột đúng
    nhưng nội dung sai.

    Không cache: sửa file rồi chạy lại là phải thấy kết quả mới ngay.
    """

    path = Path(path)
    suffix = path.suffix.lower()
    if suffix in LEGACY_SUFFIXES:
        raise StudyMappingError(
            f"'{path.name}': định dạng .xls (Excel 97-2003) không đọc được.\n"
            f"      Mở file và lưu lại dạng .xlsx rồi chọn lại."
        )
    if suffix not in SUPPORTED_SUFFIXES:
        raise StudyMappingError(
            f"'{path.name}': không phải file Excel "
            f"({', '.join(sorted(SUPPORTED_SUFFIXES))})."
        )
    if not path.is_file():
        raise StudyMappingError(f"Không tìm thấy file mapping: {path}")

    from openpyxl import load_workbook

    try:
        # data_only: file người dùng làm thường có công thức VLOOKUP.
        wb = load_workbook(path, read_only=True, data_only=True)
    except Exception as e:  # openpyxl ném đủ loại exception cho file hỏng
        raise StudyMappingError(f"'{path.name}': không mở được ({e}).") from e

    try:
        sheet = wb[wb.sheetnames[0]]
        sheet_rows = sheet.iter_rows(values_only=True)
        try:
            header_row = next(sheet_rows)
        except StopIteration:
            raise StudyMappingError(
                f"'{path.name}': sheet đầu tiên không có dòng nào."
            ) from None

        headers = ["" if c is None else str(c).strip() for c in header_row]
        index = _index_headers(headers)
        if COL_STUDY_ID not in index or COL_HRN not in index:
            return None  # không phải danh sách nghiên cứu → để mapping_excel lo

        col_study = index[COL_STUDY_ID]
        col_hrn = index[COL_HRN]
        start_column, col_start = _pick_column(index, START_DATE_HEADERS)
        end_column, col_end = _pick_column(index, END_DATE_HEADERS)

        by_hrn: Dict[str, StudyRow] = {}
        by_suffix: Dict[str, StudyRow] = {}
        ambiguous_suffixes: set = set()
        duplicates: List[str] = []
        n_missing_study_id = 0

        for row_no, raw_row in enumerate(sheet_rows, start=2):
            if raw_row is None:
                continue
            hrn = _cell_key(_cell_at(raw_row, col_hrn))
            if not hrn:
                continue  # dòng trống hoặc thiếu HRN — không phải lỗi
            study_id = _cell_display(_cell_at(raw_row, col_study))
            if not study_id:
                n_missing_study_id += 1
                continue  # có HRN nhưng không có mã nghiên cứu thì gắn gì?
            if hrn in by_hrn:
                # Giữ dòng đầu để kết quả không phụ thuộc thứ tự trong file.
                duplicates.append(hrn)
                continue

            entry = StudyRow(
                study_id=study_id,
                start_date=_parse_date_cell(
                    _cell_at(raw_row, col_start), start_column or "", row_no),
                end_date=_parse_date_cell(
                    _cell_at(raw_row, col_end), end_column or "", row_no),
            )
            by_hrn[hrn] = entry

            suffix_key = hrn[-ID_SUFFIX_LENGTH:] if len(hrn) >= ID_SUFFIX_LENGTH else None
            if suffix_key:
                if suffix_key in by_suffix and by_suffix[suffix_key] is not entry:
                    # Hai HRN khác nhau chung 10 ký tự cuối: không thể biết ID
                    # thuộc về ai, nên bỏ hẳn khoá này thay vì gán bừa.
                    ambiguous_suffixes.add(suffix_key)
                else:
                    by_suffix[suffix_key] = entry
    finally:
        wb.close()

    for key in ambiguous_suffixes:
        by_suffix.pop(key, None)
    if ambiguous_suffixes:
        log.warning(
            "%s: %d nhóm HRN trùng %d ký tự cuối — bỏ khớp theo ID cho các mã "
            "này để tránh gán nhầm: %s",
            path.name, len(ambiguous_suffixes), ID_SUFFIX_LENGTH,
            ", ".join(sorted(ambiguous_suffixes)[:_MAX_LOGGED_DUPLICATES]),
        )

    if not by_hrn:
        raise StudyMappingError(
            f"'{path.name}': không có dòng nào có đủ {COL_HRN} và {COL_STUDY_ID}."
        )

    if duplicates:
        shown = ", ".join(duplicates[:_MAX_LOGGED_DUPLICATES])
        extra = (
            f" (và {len(duplicates) - _MAX_LOGGED_DUPLICATES} mã khác)"
            if len(duplicates) > _MAX_LOGGED_DUPLICATES else ""
        )
        log.warning(
            "%s: %d dòng trùng %s, chỉ giữ dòng đầu: %s%s",
            path.name, len(duplicates), COL_HRN, shown, extra,
        )
    if n_missing_study_id:
        log.warning(
            "%s: %d dòng có %s nhưng bỏ trống %s — đã bỏ qua.",
            path.name, n_missing_study_id, COL_HRN, COL_STUDY_ID,
        )

    log.info(
        "Đã nạp danh sách nghiên cứu %s: %d %s, lọc ngày theo %s.",
        path.name, len(by_hrn), COL_HRN,
        f"{start_column or '(không có)'} → {end_column or '(không có)'}"
        if (start_column or end_column) else "(không lọc)",
    )
    return StudyMapping(
        by_hrn=MappingProxyType(by_hrn),
        by_suffix=MappingProxyType(by_suffix),
        source_name=path.name,
        start_column=start_column,
        end_column=end_column,
    )


# File BƯỚC 2 của người dùng theo quy ước tên
# ``LabRequest_<phần đuôi>.xlsx`` (ví dụ ``LabRequest_13NV-XXX_dd.mm.yyyy_.xlsx``).
INPUT_FILENAME_PREFIX = "LabRequest"
OUTPUT_FILENAME_PREFIX = "LabResult"

_INPUT_FILENAME_RE = re.compile(
    r"^" + re.escape(INPUT_FILENAME_PREFIX) + r"(_.*)?$", re.IGNORECASE
)


def derive_output_filename(mapping_filename) -> Optional[str]:
    """Suy tên file Excel xuất ra từ tên file mapping BƯỚC 2.

    File mapping theo đúng quy ước ``LabRequest_<đuôi>.xlsx`` thì phần đuôi
    (mã bệnh nhân/khoa, ngày…) được GIỮ NGUYÊN, chỉ tiền tố đổi thành
    ``LabResult`` — output phải cùng một "lô" với input để người dùng đối
    chiếu qua tên file mà không cần mở ra xem. Trả ``None`` nếu tên file
    không theo quy ước này; khi đó caller giữ nguyên tên output đang có
    (mặc định hoặc do người dùng tự chọn), không đoán bừa.
    """

    name = Path(mapping_filename).name
    stem = Path(name).stem
    ext = Path(name).suffix
    m = _INPUT_FILENAME_RE.match(stem)
    if not m:
        return None
    suffix = m.group(1) or ""
    return f"{OUTPUT_FILENAME_PREFIX}{suffix}{ext}"


__all__ = [
    "COL_HRN",
    "COL_STUDY_ID",
    "END_DATE_HEADERS",
    "ID_SUFFIX_LENGTH",
    "INPUT_FILENAME_PREFIX",
    "OUTPUT_FILENAME_PREFIX",
    "START_DATE_HEADERS",
    "StudyMapping",
    "StudyMappingError",
    "StudyRow",
    "derive_output_filename",
    "load_study_mapping",
    "parse_record_date",
]
