"""File Excel mapping do NGƯỜI DÙNG cung cấp — gắn thêm cột vào bản ghi XML4.

Đây là loại mapping thứ hai, KHÁC HẲN ``medical_catalog`` — đừng bê cơ chế bên
đó sang đây:

===================  ==========================  =============================
                     medical_catalog             mapping_excel (file này)
===================  ==========================  =============================
Nguồn                dữ liệu của bản phát hành   người dùng tự chọn mỗi lần chạy
Vân tay SHA-256      có, sai là dừng hẳn         KHÔNG (file của người dùng)
Cache theo process   có                          KHÔNG — sửa file, chạy lại là
                                                 phải thấy kết quả mới ngay
Bắt buộc             có, thiếu là hỏng cả run    tuỳ chọn, bỏ trống vẫn chạy
Vai trò              lọc Include/Exclude         chỉ gắn THÊM cột, không lọc
===================  ==========================  =============================

Quy ước file (người dùng kiểm soát hoàn toàn, không hardcode trong code):

- Dòng đầu tiên là tiêu đề cột.
- **Cột đầu tiên là khoá join**: tên của nó phải trùng tên một trường trong
  bản ghi XML4 (``MA_DICH_VU``, ``MA_LK``, ``MA_BENH``…). Cố ý làm vậy để đổi
  tiêu chí join chỉ cần sửa tiêu đề trong Excel, không phải sửa code.
- Các cột còn lại được gắn thêm vào bản ghi nào khớp khoá.

Mapping **không bao giờ ghi đè** dữ liệu sẵn có của bản ghi: cột trùng tên sẽ
được đổi thành ``<tên>_MAP``. Ghi đè âm thầm lên ``Name_Method`` (do danh mục
điền) hay lên chính trường gốc trong XML là kiểu mất dữ liệu không ai phát
hiện ra.
"""

from __future__ import annotations

import logging
from dataclasses import dataclass
from datetime import date, datetime, time
from pathlib import Path
from types import MappingProxyType
from typing import Dict, List, Mapping, Tuple

log = logging.getLogger(__name__)

# openpyxl chỉ đọc được .xlsx/.xlsm. File .xls (Excel 97-2003) cần thư viện
# khác (xlrd) — không có trong requirements, nên báo lỗi rõ thay vì để openpyxl
# ném ra một traceback khó hiểu.
SUPPORTED_SUFFIXES = frozenset({".xlsx", ".xlsm"})
LEGACY_SUFFIXES = frozenset({".xls"})

# Hậu tố cho cột mapping bị trùng tên với cột đã có trong bản ghi.
COLLISION_SUFFIX = "_MAP"

# Số dòng trùng khoá tối đa ghi ra log (tránh log dài vô ích).
_MAX_LOGGED_DUPLICATES = 10


class MappingError(Exception):
    """Không dùng được file mapping. Message viết cho người dùng cuối đọc."""


@dataclass(frozen=True)
class ExcelMapping:
    """Nội dung file mapping đã đọc xong.

    ``rows`` là ``{khoá đã chuẩn hoá: {tên cột: giá trị}}``; tên cột trong đó
    đã được xử lý trùng lặp, dùng thẳng làm tên cột đầu ra được.
    """

    key_column: str
    value_columns: Tuple[str, ...]
    rows: Mapping[str, Mapping[str, str]]
    source_name: str

    def lookup(self, value) -> Mapping[str, str] | None:
        """Tra 1 giá trị khoá. ``None`` = không có dòng nào khớp."""

        key = normalize_key(value)
        if not key:
            return None
        return self.rows.get(key)

    def __len__(self) -> int:
        return len(self.rows)


def normalize_key(value) -> str:
    """Chuẩn hoá khoá join: bỏ khoảng trắng, không phân biệt hoa/thường.

    Cùng quy tắc với ``medical_catalog.normalize_service_code`` (mã dịch vụ có
    cả chữ thường, ví dụ ``27.205b.0463``), nhưng cố tình KHÔNG dùng chung hàm:
    hai file mapping độc lập nhau, đổi quy tắc bên này không được kéo theo bên
    kia.
    """

    return str(value or "").strip().upper()


def _cell_text(value) -> str:
    """Ép giá trị ô Excel về chuỗi giống cách bản ghi XML4 lưu dữ liệu.

    openpyxl trả số nguyên dưới dạng float (``1054.0``) — để nguyên thì cột mã
    trong Excel sẽ không bao giờ khớp với chuỗi ``"1054"`` đọc từ XML.
    """

    if value is None:
        return ""
    if isinstance(value, bool):
        return "1" if value else "0"
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    if isinstance(value, datetime):
        return value.strftime("%d/%m/%Y %H:%M:%S") if (
            value.hour or value.minute or value.second
        ) else value.strftime("%d/%m/%Y")
    if isinstance(value, date):
        return value.strftime("%d/%m/%Y")
    if isinstance(value, time):
        return value.strftime("%H:%M:%S")
    return str(value).strip()


def _resolve_value_columns(headers: List[str], reserved) -> Tuple[str, ...]:
    """Đặt tên cột đầu ra, tránh trùng với cột đã có và trùng lẫn nhau."""

    taken = set(reserved)
    resolved: List[str] = []
    for name in headers:
        candidate = name
        while candidate in taken:
            candidate += COLLISION_SUFFIX
        if candidate != name:
            log.warning(
                "Cột mapping %r trùng tên cột đã có — đổi thành %r.",
                name, candidate,
            )
        taken.add(candidate)
        resolved.append(candidate)
    return tuple(resolved)


def _read_header(sheet_rows, path: Path) -> List[str]:
    """Lấy dòng tiêu đề. Ném ``MappingError`` nếu không dùng được."""

    try:
        header_row = next(sheet_rows)
    except StopIteration:
        raise MappingError(f"'{path.name}': sheet đầu tiên không có dòng nào.") from None

    headers = [_cell_text(c) for c in header_row]
    while headers and not headers[-1]:
        headers.pop()  # bỏ các cột rỗng ở đuôi do Excel hay để lại

    if not headers or not headers[0]:
        raise MappingError(
            f"'{path.name}': ô đầu tiên (A1) đang trống.\n"
            f"      A1 phải là TÊN TRƯỜNG trong hồ sơ XML4 dùng để ghép dữ "
            f"liệu, ví dụ MA_DICH_VU."
        )
    if len(headers) < 2:
        raise MappingError(
            f"'{path.name}': chỉ có 1 cột ({headers[0]}).\n"
            f"      Cần thêm ít nhất 1 cột nữa — đó mới là phần được gắn vào "
            f"kết quả."
        )
    return headers


def load_mapping(path, reserved_columns=()) -> ExcelMapping:
    """Đọc file Excel mapping. Ném ``MappingError`` với message tiếng Việt.

    ``reserved_columns`` là các tên cột đã có sẵn trong bản ghi (để tránh ghi
    đè). Không cache: người dùng sửa file rồi chạy lại phải thấy ngay kết quả
    mới.
    """

    path = Path(path)
    suffix = path.suffix.lower()
    if suffix in LEGACY_SUFFIXES:
        raise MappingError(
            f"'{path.name}': định dạng .xls (Excel 97-2003) không đọc được.\n"
            f"      Mở file và lưu lại dạng .xlsx rồi chọn lại."
        )
    if suffix not in SUPPORTED_SUFFIXES:
        raise MappingError(
            f"'{path.name}': không phải file Excel "
            f"({', '.join(sorted(SUPPORTED_SUFFIXES))})."
        )
    if not path.is_file():
        raise MappingError(f"Không tìm thấy file mapping: {path}")

    from openpyxl import load_workbook

    try:
        # data_only: lấy giá trị đã tính sẵn thay vì công thức, vì file do
        # người dùng làm thường có công thức VLOOKUP.
        wb = load_workbook(path, read_only=True, data_only=True)
    except Exception as e:  # openpyxl ném đủ loại exception cho file hỏng
        raise MappingError(f"'{path.name}': không mở được ({e}).") from e

    try:
        sheet = wb[wb.sheetnames[0]]
        sheet_rows = sheet.iter_rows(values_only=True)
        headers = _read_header(sheet_rows, path)
        key_column = headers[0]
        value_columns = _resolve_value_columns(headers[1:], reserved_columns)

        rows: Dict[str, Dict[str, str]] = {}
        duplicates: List[str] = []
        for raw_row in sheet_rows:
            key = normalize_key(raw_row[0] if raw_row else None)
            if not key:
                continue  # dòng trống hoặc thiếu khoá — bỏ qua, không phải lỗi
            if key in rows:
                # Giữ dòng đầu để kết quả không phụ thuộc thứ tự trong file.
                duplicates.append(key)
                continue
            rows[key] = {
                column: _cell_text(raw_row[i + 1]) if i + 1 < len(raw_row) else ""
                for i, column in enumerate(value_columns)
            }
    finally:
        wb.close()

    if not rows:
        raise MappingError(
            f"'{path.name}': không có dòng dữ liệu nào ở cột {key_column}."
        )

    if duplicates:
        shown = ", ".join(duplicates[:_MAX_LOGGED_DUPLICATES])
        extra = f" (và {len(duplicates) - _MAX_LOGGED_DUPLICATES} khoá khác)" if (
            len(duplicates) > _MAX_LOGGED_DUPLICATES
        ) else ""
        log.warning(
            "%s: %d dòng trùng khoá %s, chỉ giữ dòng đầu: %s%s",
            path.name, len(duplicates), key_column, shown, extra,
        )

    log.info(
        "Đã nạp mapping %s: %d dòng, join theo %s, thêm cột: %s",
        path.name, len(rows), key_column, ", ".join(value_columns),
    )
    return ExcelMapping(
        key_column=key_column,
        value_columns=value_columns,
        rows=MappingProxyType(rows),
        source_name=path.name,
    )


__all__ = [
    "COLLISION_SUFFIX",
    "LEGACY_SUFFIXES",
    "SUPPORTED_SUFFIXES",
    "ExcelMapping",
    "MappingError",
    "load_mapping",
    "normalize_key",
]
