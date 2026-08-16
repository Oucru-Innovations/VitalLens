"""Xử lý XML → Excel: giải mã Base64 các FILEHOSO loại XML4, xuất Excel.

Chỉ XML4 được xử lý; các loại hồ sơ khác (XML1/2/3/…) bị bỏ qua hoàn toàn.

Mỗi bản ghi được đối chiếu với danh mục dịch vụ (``database_medical.csv``)
qua ``MA_DICH_VU`` = ``ID_SERVICE`` để lấy tên phương pháp (cột ``Name_Method``
trong danh mục, xuất ra Excel dưới tên ``TEN_DICH_VU``) và nhóm lọc:

    Include                   -> sheet chính
    Exclude                   -> loại khỏi kết quả
    (không có trong danh mục)  -> sheet riêng để rà soát thủ công

Bản ghi thiếu ``MA_DICH_VU`` cũng vào sheet "chưa phân loại" — không có mã thì
không thể khẳng định là Exclude, nên không được lặng lẽ vứt đi.

Ngoài danh mục cố định đó, người dùng có thể chọn thêm 1 file Excel ở BƯỚC 2.
File này tuỳ chọn và được nhận dạng theo tiêu đề cột, chạy một trong hai chế độ:

- Có đủ cột ``USUBJID`` + ``EMR_ID`` → **danh sách nghiên cứu**
  (``services/study_mapping.py``): gắn mã nghiên cứu (xuất ra dưới tên cột
  ``MA_NTG``), lọc theo khoảng ngày (bản ghi ngoài khoảng bị LOẠI HẲN, không
  vào sheet nào), và ẩn ``ID``/``MA_LK`` khỏi file xuất ra. Chế độ này còn thêm
  sheet ``Summary`` tổng hợp theo ``MA_NTG`` và, nếu tên file mapping theo quy
  ước ``LabRequest_<đuôi>.xlsx``,
  gợi ý luôn tên file xuất ra ``LabResult_<đuôi>.xlsx``
  (``study_mapping.derive_output_filename``, dùng ở ``pages/xml_page.py``).
- Ngược lại → **mapping tổng quát** (``services/mapping_excel.py``): ô A1 là tên
  trường XML4 dùng để ghép, các cột còn lại được gắn thêm, không lọc gì.

Payload giải mã/parse không được thì bị BỎ, không xuất nội dung thô ra Excel:
bộ lọc PII ở đây chạy trên tên cột (``XML4_EXCLUDED_COLUMNS`` trong
``_sheet_columns``) nên nó không lọc nổi một ô chứa nguyên văn hồ sơ. Số hồ sơ
và số file bị bỏ được đếm và báo lên UI.
"""

import os
import logging
import base64
import zlib
import xml.etree.ElementTree as ET
from concurrent.futures import ThreadPoolExecutor, as_completed

from apps.services import mapping_excel, medical_catalog, study_mapping
from apps.services.excel_export import (
    append_row_as_text,
    collect_columns,
    sanitize_sheet_name,
)

log = logging.getLogger(__name__)

# Loại hồ sơ duy nhất được xuất ra Excel.
TARGET_LOAIHOSO = "XML4"

# XML4: lấy hết các cột, trừ những cột này.
XML4_EXCLUDED_COLUMNS = {"MA_BS_DOC_KQ"}

# Cột mã dịch vụ trong XML4.
SERVICE_CODE_COLUMN = "MA_DICH_VU"
# Tên cũ là medical_catalog.COL_NAME, nhưng output sẽ là ``TEN_DICH_VU`` để đồng bộ cách đặt tên
NAME_METHOD_COLUMN = "TEN_DICH_VU"    # giá trị lấy từ medical_catalog.COL_NAME

# Cột dùng cho danh sách nghiên cứu (BƯỚC 2, chế độ USUBJID/EMR_ID).
STUDY_ID_COLUMN = "MA_NTG"            # giá trị lấy từ study_mapping.COL_STUDY_ID
LINK_CODE_COLUMN = "MA_LK"          # khớp TOÀN BỘ với EMR_ID
RESULT_DATE_COLUMN = "NGAY_KQ"      # yyyymmddhhmm — nguồn để lọc theo khoảng ngày

# Các trường thử khớp 10 ký tự cuối với EMR_ID, theo đúng thứ tự ưu tiên. ``ID``
# là tên file XML (process_xml_file ghi đè), ``ID_GOC`` là ID gốc trong hồ sơ.
STUDY_SUFFIX_FIELDS = ("ID", "ID_GOC")

# Đã có MA_NTG thì hai định danh gốc này bị ẩn khỏi file xuất ra (yêu cầu ẩn
# danh). ID_GOC được GIỮ LẠI có chủ đích — người dùng cần nó để đối chiếu.
STUDY_HIDDEN_COLUMNS = frozenset({"ID", LINK_CODE_COLUMN})

# Tên sheet (ASCII: sanitize_sheet_name thay dấu tiếng Việt bằng "_").
SHEET_INCLUDE = "XML4_Include"
SHEET_UNCLASSIFIED = "XML4_ChuaPhanLoai"
SHEET_NO_STUDY = "XML4_KhongKhopHRN"
SHEET_SUMMARY = "Summary"

# Cột cố định của sheet Summary (chỉ có khi chạy chế độ danh sách nghiên cứu).
SUMMARY_COLUMNS = (
    "Mapping_TuNgay", "Mapping_DenNgay",
    "XML_TuNgay", "XML_DenNgay",
    "So_MaDichVu", "So_Row_HopLe", "So_Row_Unknown",
)

# Hai byte đầu của một stream GZIP.
GZIP_MAGIC = b"\x1f\x8b"

# Trần dung lượng sau khi giải nén 1 payload. Một hồ sơ XML4 thật chỉ vài trăm
# KB; vượt xa mức này thì là file hỏng hoặc "gzip bomb" (vài KB base64 nở ra
# hàng GB). Không chặn thì 8 luồng chạy song song đủ làm app hết RAM.
MAX_DECOMPRESSED_BYTES = 64 * 1024 * 1024

# Số tên file tối đa liệt kê trong cảnh báo (tránh hộp thoại dài vô tận).
MAX_LISTED_FAILED_FILES = 5


def _gunzip_bounded(data):
    """Giải nén GZIP có trần dung lượng. Trả ``None`` nếu hỏng hoặc quá lớn."""

    dobj = zlib.decompressobj(wbits=16 + zlib.MAX_WBITS)  # 16 = đọc header GZIP
    try:
        out = dobj.decompress(data, MAX_DECOMPRESSED_BYTES)
    except zlib.error as e:
        log.warning("Stream GZIP hỏng, bỏ qua payload: %s", e)
        return None

    if dobj.unconsumed_tail or not dobj.eof:
        log.warning(
            "Payload GZIP vượt trần %d byte hoặc bị cắt cụt — bỏ qua.",
            MAX_DECOMPRESSED_BYTES,
        )
        return None
    return out


def decode_base64_content(b64_string):
    try:
        decoded_bytes = base64.b64decode(b64_string)
    except Exception as e:
        log.warning("Giải mã Base64 thất bại: %s", e)
        return None

    # Không có magic GZIP → nội dung để trần, dùng nguyên bytes đã giải Base64.
    if decoded_bytes.startswith(GZIP_MAGIC):
        decoded_bytes = _gunzip_bounded(decoded_bytes)
        if decoded_bytes is None:
            return None

    try:
        return decoded_bytes.decode("utf-8-sig")
    except UnicodeDecodeError as e:
        log.warning("Giải mã văn bản thất bại sau Base64/GZIP: %s", e)
        return None


def parse_inner_xml(xml_string):
    """Parse XML con đã giải mã. Trả ``None`` nếu không parse được.

    KHÔNG bao giờ trả nội dung thô ra ngoài. Bản cũ nhét cả payload vào một ô
    ``RAW_CONTENT`` khi parse lỗi; bản ghi đó không có ``MA_DICH_VU`` nên rơi
    vào sheet "chưa phân loại" và mang theo NGUYÊN VĂN hồ sơ — gồm cả các
    trường trong ``XML4_EXCLUDED_COLUMNS``. Bộ lọc cột chỉ chạy ở
    ``_sheet_columns()``, tức là chạy trên TÊN cột, nên nó không thể lọc được
    thứ đã bị gộp vào trong một ô. Payload hỏng nay chỉ được đếm và báo lên UI.
    """

    try:
        return _extract_all_records(ET.fromstring(xml_string.strip()))
    except ET.ParseError as e:
        log.warning("Payload XML4 không parse được, bỏ qua: %s", e)
        return None


def _extract_all_records(root):
    records = []
    list_parents = []
    for child in root:
        grandchildren = list(child)
        if grandchildren:
            if any(list(gc) for gc in grandchildren):
                list_parents.append(child)

    for list_parent in list_parents:
        for item in list_parent:
            sub_children = list(item)
            if sub_children:
                record = {}
                for field in sub_children:
                    field_children = list(field)
                    if not field_children:
                        text = (field.text or "").strip()
                        record[field.tag] = text
                    else:
                        texts = [s.text.strip() for s in field.iter()
                                 if s.text and s.text.strip()]
                        record[field.tag] = "; ".join(texts)
                if record:
                    records.append(record)

    if not records:
        record = {}
        for child in root:
            sub_children = list(child)
            if not sub_children:
                record[child.tag] = (child.text or "").strip()
            else:
                for gc in sub_children:
                    if not list(gc):
                        record[f"{child.tag}_{gc.tag}"] = (gc.text or "").strip()
        if record:
            records.append(record)
    return records


def process_xml_file(xml_filepath):
    """Giải mã mọi FILEHOSO loại XML4 trong 1 file.

    Trả ``(records, n_payload_loi)`` — ``n_payload_loi`` là số payload XML4
    giải mã/parse không được. Phải trả về để ``_collect_and_save`` còn báo cho
    người dùng biết có bao nhiêu hồ sơ đã bị bỏ, thay vì chỉ ghi vào log.
    """

    filename = os.path.basename(xml_filepath)
    records_out = []
    n_bad = 0

    try:
        tree = ET.parse(xml_filepath)
        root = tree.getroot()
    except ET.ParseError:
        with open(xml_filepath, "r", encoding="utf-8-sig") as f:
            content = f.read()
        root = ET.fromstring(content)

    file_hoso_list = root.findall(".//FILEHOSO")
    for fh in file_hoso_list:
        loai_node = fh.find("LOAIHOSO")
        content_node = fh.find("NOIDUNGFILE")
        if loai_node is None or content_node is None:
            continue
        loai_hoso = (loai_node.text or "").strip()
        if loai_hoso != TARGET_LOAIHOSO:
            continue
        b64_content = content_node.text.strip() if content_node.text else ""
        if not b64_content:
            continue
        decoded_xml = decode_base64_content(b64_content)
        if decoded_xml is None:
            n_bad += 1
            continue
        records = parse_inner_xml(decoded_xml)
        if records is None:
            n_bad += 1
            continue
        file_id = os.path.splitext(filename)[0]
        for rec in records:
            if "ID" in rec:
                rec["ID_GOC"] = rec["ID"]
            rec["ID"] = file_id
        records_out.extend(records)

    return records_out, n_bad


def _sheet_columns(records, extra_columns=(), lead_column="ID", hidden_columns=()):
    """Cột của 1 sheet: ``lead_column`` đứng đầu, TEN_DICH_VU ngay sau MA_DICH_VU.

    ``extra_columns`` (các cột từ file mapping tổng quát) luôn được đưa vào
    cuối, kể cả khi không bản ghi nào khớp — người dùng đã chọn file mapping thì
    phải thấy cột của nó, dù rỗng, mới biết là không khớp được gì.

    ``lead_column`` / ``hidden_columns`` khác nhau theo từng sheet: sheet có
    MA_NTG thì ẩn ID và MA_LK, còn sheet "không khớp EMR_ID" phải giữ nguyên
    hai cột đó — nó tồn tại để người dùng rà soát, ẩn định danh đi thì rà bằng
    gì.
    """

    columns = [lead_column] + [
        k for k in collect_columns(records)
        if k not in (lead_column, NAME_METHOD_COLUMN)
        and k not in XML4_EXCLUDED_COLUMNS
        and k not in hidden_columns
        and k not in extra_columns
    ]
    if SERVICE_CODE_COLUMN in columns:
        columns.insert(columns.index(SERVICE_CODE_COLUMN) + 1, NAME_METHOD_COLUMN)
    else:
        columns.append(NAME_METHOD_COLUMN)
    return columns + list(extra_columns)


def _split_by_catalog(records, catalog):
    """Chia bản ghi theo nhóm trong danh mục.

    Trả ``(included, unclassified, n_excluded)``. Group lạ (không phải
    Include/Exclude) được xếp vào nhóm chưa phân loại thay vì bị loại — sai
    chính tả trong danh mục không được âm thầm làm mất dữ liệu. Việc cảnh báo
    Group lạ do ``medical_catalog`` lo một lần lúc nạp file, không lặp lại ở đây.

    Lưu ý: hàm SỬA TRỰC TIẾP các dict trong ``records`` (gắn thêm cột
    ``TEN_DICH_VU``) thay vì copy — bản ghi có thể tới hàng trăm nghìn dòng,
    nhân đôi lên chỉ để giữ tính "thuần" là không đáng. Bản ghi thuộc nhóm
    Exclude không được gắn gì vì đằng nào cũng bị bỏ.
    """

    included, unclassified, n_excluded = [], [], 0
    for rec in records:
        entry = medical_catalog.lookup(catalog, rec.get(SERVICE_CODE_COLUMN))
        if entry is None:
            rec[NAME_METHOD_COLUMN] = ""
            unclassified.append(rec)
        elif entry.group == medical_catalog.GROUP_EXCLUDE:
            n_excluded += 1
        elif entry.group == medical_catalog.GROUP_INCLUDE:
            rec[NAME_METHOD_COLUMN] = entry.name_method
            included.append(rec)
        else:
            rec[NAME_METHOD_COLUMN] = entry.name_method
            unclassified.append(rec)
    return included, unclassified, n_excluded


def _apply_mapping(records, mapping):
    """Gắn các cột từ file mapping vào bản ghi khớp khoá. Trả số bản ghi khớp.

    Bản ghi không khớp vẫn được điền chuỗi rỗng cho đủ cột — để trống hẳn thì
    ``rec.get(c, "")`` lúc ghi Excel cũng ra rỗng, nhưng điền sẵn giúp
    ``collect_columns`` nhìn thấy cột và giữ thứ tự ổn định.

    Sửa trực tiếp dict trong ``records``, cùng lý do với ``_split_by_catalog``.
    """

    n_matched = 0
    blank = {column: "" for column in mapping.value_columns}
    for rec in records:
        found = mapping.lookup(rec.get(mapping.key_column))
        if found is None:
            rec.update(blank)
        else:
            rec.update(found)
            n_matched += 1
    return n_matched


def _match_study(rec, study):
    """Tìm dòng danh sách nghiên cứu ứng với 1 bản ghi XML4.

    Thứ tự thử: ``MA_LK`` khớp TOÀN BỘ EMR_ID trước, sau đó ``ID`` rồi ``ID_GOC``
    khớp 10 ký tự cuối. Khớp toàn bộ chắc chắn hơn nên phải được ưu tiên; khớp
    theo hậu tố chỉ là phương án dự phòng khi hồ sơ không có MA_LK.
    """

    row = study.lookup_full(rec.get(LINK_CODE_COLUMN))
    if row is not None:
        return row
    for field in STUDY_SUFFIX_FIELDS:
        row = study.lookup_suffix(rec.get(field))
        if row is not None:
            return row
    return None


def _new_study_stats():
    return {"matched": 0, "no_hrn": 0, "out_of_range": 0, "missing_date": 0}


def _split_by_study(records, study, stats):
    """Chia bản ghi theo danh sách nghiên cứu. Trả ``(khớp, không khớp EMR_ID)``.

    Bản ghi khớp ``EMR_ID`` được gắn ``MA_NTG``. Bản ghi KHÔNG khớp EMR_ID rơi
    vào nhóm "không khớp" thứ hai — nhóm này vẫn được xuất ra sheet riêng
    (``XML4_KhongKhopHRN``) để người dùng rà lại file danh sách của mình.

    Bản ghi khớp EMR_ID nhưng ``NGAY_KQ`` nằm NGOÀI khoảng ngày của mapping thì
    bị LOẠI HẲN — không vào ``matched`` lẫn ``unmatched``, tức không được ghi
    vào bất kỳ sheet nào. Khác với việc không khớp EMR_ID (có thể là lỗi đánh
    máy trong file danh sách, cần rà lại), nằm ngoài khoảng ngày là bằng chứng
    RÕ RÀNG bản ghi không thuộc phạm vi nghiên cứu, nên không có gì để rà soát.

    Bản ghi khớp EMR_ID nhưng thiếu/hỏng ``NGAY_KQ`` thì được GIỮ (vào
    ``matched``): không có ngày không phải là bằng chứng nằm ngoài khoảng
    nghiên cứu (cùng lối xử lý với ``MA_DICH_VU`` rỗng ở ``_split_by_catalog``).
    Số này được đếm riêng và báo lên UI để người dùng tự quyết.

    Sửa trực tiếp dict trong ``records``, cùng lý do với ``_split_by_catalog``.
    """

    matched, unmatched = [], []
    for rec in records:
        row = _match_study(rec, study)
        if row is None:
            stats["no_hrn"] += 1
            unmatched.append(rec)
            continue

        if row.has_range:
            day = study_mapping.parse_record_date(rec.get(RESULT_DATE_COLUMN))
            if day is None:
                stats["missing_date"] += 1
            elif not row.covers(day):
                stats["out_of_range"] += 1
                continue  # loại hẳn — không vào sheet nào

        rec[STUDY_ID_COLUMN] = row.study_id
        stats["matched"] += 1
        matched.append(rec)
    return matched, unmatched


def _study_note(study, stats):
    """Câu tóm tắt kết quả ghép danh sách nghiên cứu, hiện lên UI."""

    parts = [
        f"\nDanh sách nghiên cứu '{study.source_name}': "
        f"{stats['matched']} bản ghi có {STUDY_ID_COLUMN}"
    ]
    if study.has_date_filter:
        parts.append(
            f", lọc {RESULT_DATE_COLUMN} theo "
            f"{study.start_column or '(không có)'} → "
            f"{study.end_column or '(không có)'} (trọn ngày)"
        )
    parts.append(".")
    if stats["no_hrn"]:
        parts.append(
            f" {stats['no_hrn']} không khớp {study_mapping.COL_HRN} "
            f"→ sheet {SHEET_NO_STUDY}."
        )
    if stats["out_of_range"]:
        parts.append(
            f"{stats['out_of_range']} bản ghi của {study_mapping.COL_HRN} với {RESULT_DATE_COLUMN} nằm ngoài khung thời gian sẽ không được thu thập."
        )
    if stats["missing_date"]:
        parts.append(
            f" ⚠ {stats['missing_date']} bản ghi khớp {study_mapping.COL_HRN} "
            f"nhưng thiếu {RESULT_DATE_COLUMN} — đã giữ lại, cần rà thủ công."
        )
    return "".join(parts)


def _fmt_yyyymmdd(day):
    """``19850102`` -> ``"1985-01-02"``. ``None`` -> chuỗi rỗng."""

    if day is None:
        return ""
    return f"{day // 10000:04d}-{day // 100 % 100:02d}-{day % 100:02d}"


def _study_mapping_ranges(study):
    """``MA_NTG`` -> (ngày bắt đầu nhỏ nhất, ngày kết thúc lớn nhất) theo file mapping.

    Mỗi mã nghiên cứu có thể lặp lại trên nhiều dòng (một dòng/bệnh nhân) với
    khoảng ngày riêng từng dòng; sheet Summary báo khoảng NGOÀI CÙNG bao trọn
    mọi bệnh nhân của nghiên cứu đó, không phải khoảng của một dòng bất kỳ.
    """

    ranges = {}
    for row in study.by_hrn.values():
        start, end = ranges.setdefault(row.study_id, [None, None])
        if row.start_date is not None:
            start = row.start_date if start is None else min(start, row.start_date)
        if row.end_date is not None:
            end = row.end_date if end is None else max(end, row.end_date)
        ranges[row.study_id] = [start, end]
    return ranges


def _build_summary_rows(included, unclassified, study):
    """Gộp thống kê theo ``MA_NTG`` cho sheet Summary.

    So sánh khoảng ngày KHAI BÁO trong file mapping với khoảng ngày THỰC TẾ có
    trong dữ liệu XML đã xuất — hai khoảng lệch nhau là dấu hiệu người dùng
    nên xem lại (ví dụ nghiên cứu khai 6 tháng nhưng dữ liệu XML chỉ có 2
    tháng đầu). Chỉ tính trên bản ghi ĐÃ xuất ra sheet (khớp EMR_ID và, nếu có
    khoảng ngày, nằm trong khoảng) — bản ghi bị loại ở ``_split_by_study``
    không góp phần vào đây, cùng logic với việc chúng không vào sheet nào.
    """

    mapping_ranges = _study_mapping_ranges(study)
    stats = {}
    for records, is_valid in ((included, True), (unclassified, False)):
        for rec in records:
            study_id = rec.get(STUDY_ID_COLUMN)
            if not study_id:
                continue
            entry = stats.setdefault(study_id, {
                "n_valid": 0, "n_unknown": 0, "services": set(),
                "xml_min": None, "xml_max": None,
            })
            entry["n_valid" if is_valid else "n_unknown"] += 1
            code = rec.get(SERVICE_CODE_COLUMN)
            if code:
                entry["services"].add(code)
            day = study_mapping.parse_record_date(rec.get(RESULT_DATE_COLUMN))
            if day is not None:
                entry["xml_min"] = day if entry["xml_min"] is None else min(entry["xml_min"], day)
                entry["xml_max"] = day if entry["xml_max"] is None else max(entry["xml_max"], day)

    rows = []
    for study_id in sorted(stats):
        entry = stats[study_id]
        map_start, map_end = mapping_ranges.get(study_id, (None, None))
        rows.append({
            STUDY_ID_COLUMN: study_id,
            "Mapping_TuNgay": _fmt_yyyymmdd(map_start),
            "Mapping_DenNgay": _fmt_yyyymmdd(map_end),
            "XML_TuNgay": _fmt_yyyymmdd(entry["xml_min"]),
            "XML_DenNgay": _fmt_yyyymmdd(entry["xml_max"]),
            "So_MaDichVu": len(entry["services"]),
            "So_Row_HopLe": entry["n_valid"],
            "So_Row_Unknown": entry["n_unknown"],
        })
    return rows


def _dropped_input_warning(failed_files, n_bad_payloads):
    """Câu cảnh báo về phần dữ liệu đầu vào đã bị bỏ. Rỗng nếu không mất gì.

    Với dữ liệu y tế, "đọc được bao nhiêu" quan trọng ngang "xuất ra bao
    nhiêu": người dùng phải biết ngay trên UI là có file/hồ sơ bị bỏ, chứ
    không phải đi mò trong log.
    """

    parts = []
    if failed_files:
        listed = ", ".join(failed_files[:MAX_LISTED_FAILED_FILES])
        if len(failed_files) > MAX_LISTED_FAILED_FILES:
            listed += f", …(+{len(failed_files) - MAX_LISTED_FAILED_FILES})"
        parts.append(f"{len(failed_files)} file không đọc được: {listed}")
    if n_bad_payloads:
        parts.append(f"{n_bad_payloads} hồ sơ {TARGET_LOAIHOSO} hỏng đã bị bỏ")
    if not parts:
        return ""
    return "\n⚠ " + "; ".join(parts) + ". Xem log để biết chi tiết."


def _collect_and_save(xml_files, output_path, mapping_path=None):
    from openpyxl import Workbook

    if not xml_files:
        return False, "Chưa chọn file XML nào!"

    # Nạp danh mục TRƯỚC khi giải mã: thiếu nó thì không lọc được, và xuất ra
    # file chưa lọc còn nguy hiểm hơn là báo lỗi.
    try:
        catalog = medical_catalog.load_catalog()
    except medical_catalog.CatalogError as e:
        return False, str(e)

    # Mapping cũng nạp trước khi giải mã: file mapping sai thì báo ngay, thay
    # vì bắt người dùng chờ hết mấy phút decode rồi mới đổ lỗi.
    #
    # Chế độ được chọn theo TIÊU ĐỀ CỘT, không theo một nút bấm riêng: file có
    # đủ USUBJID + EMR_ID là danh sách nghiên cứu, còn lại là mapping tổng quát.
    # Người dùng chỉ có một nút "Chọn file mapping" và không phải nhớ mình đang
    # ở chế độ nào.
    study = None
    mapping = None
    if mapping_path:
        try:
            study = study_mapping.load_study_mapping(mapping_path)
        except study_mapping.StudyMappingError as e:
            return False, f"Lỗi file danh sách nghiên cứu — {e}"
        if study is None:
            try:
                mapping = mapping_excel.load_mapping(
                    mapping_path,
                    reserved_columns=(
                        {"ID", "ID_GOC", STUDY_ID_COLUMN, NAME_METHOD_COLUMN}
                        | XML4_EXCLUDED_COLUMNS
                    ),
                )
            except mapping_excel.MappingError as e:
                return False, f"Lỗi file mapping — {e}"

    all_records = []
    failed_files = []
    n_bad_payloads = 0
    max_workers = min(8, len(xml_files), os.cpu_count() or 4)
    with ThreadPoolExecutor(max_workers=max_workers) as pool:
        future_to_file = {pool.submit(process_xml_file, xf): xf for xf in xml_files}
        for future in as_completed(future_to_file):
            xf = future_to_file[future]
            try:
                records, n_bad = future.result()
            except Exception as e:
                # Đếm để còn báo lên UI: chỉ ghi log thì người dùng thấy
                # "Hoàn thành!" mà không biết 1/5 số file đã bị bỏ qua.
                log.warning("Lỗi xử lý %s: %s", xf, e)
                failed_files.append(os.path.basename(xf))
                continue
            all_records.extend(records)
            n_bad_payloads += n_bad

    warning = _dropped_input_warning(failed_files, n_bad_payloads)

    if not all_records:
        return False, (
            f"Không có dữ liệu {TARGET_LOAIHOSO} trong các file đã chọn!{warning}"
        )

    included, unclassified, n_excluded = _split_by_catalog(all_records, catalog)
    if not included and not unclassified:
        return False, (
            f"Toàn bộ {n_excluded} bản ghi đều thuộc nhóm Exclude — "
            f"không còn gì để xuất.{warning}"
        )

    mapping_columns = ()
    mapping_note = ""
    if mapping is not None:
        mapping_columns = mapping.value_columns
        n_matched = (
            _apply_mapping(included, mapping) + _apply_mapping(unclassified, mapping)
        )
        mapping_note = (
            f"\nMapping '{mapping.source_name}' (join theo {mapping.key_column}): "
            f"{n_matched}/{len(included) + len(unclassified)} bản ghi khớp."
        )

    # Mỗi sheet mang bộ cột riêng: (tên sheet, bản ghi, cột đứng đầu, cột bị ẩn).
    summary_rows = []
    if study is None:
        sheets = [
            (SHEET_INCLUDE, included, "ID", ()),
            (SHEET_UNCLASSIFIED, unclassified, "ID", ()),
        ]
    else:
        stats = _new_study_stats()
        included, no_study_inc = _split_by_study(included, study, stats)
        unclassified, no_study_unc = _split_by_study(unclassified, study, stats)
        mapping_note = _study_note(study, stats)
        # Tính Summary TRƯỚC khi included/unclassified có thêm cột mapping ở
        # dưới — không ảnh hưởng gì (chỉ đọc STUDY_ID_COLUMN/SERVICE_CODE_COLUMN
        # /RESULT_DATE_COLUMN) nhưng để cạnh _split_by_study cho dễ theo dõi.
        summary_rows = _build_summary_rows(included, unclassified, study)
        sheets = [
            (SHEET_INCLUDE, included, STUDY_ID_COLUMN, STUDY_HIDDEN_COLUMNS),
            (SHEET_UNCLASSIFIED, unclassified, STUDY_ID_COLUMN, STUDY_HIDDEN_COLUMNS),
            # Sheet rà soát: giữ nguyên ID/MA_LK để còn truy ngược được.
            (SHEET_NO_STUDY, no_study_inc + no_study_unc, "ID", ()),
        ]

    # write_only: openpyxl ghi thẳng từng dòng ra file thay vì dựng toàn bộ đối
    # tượng Cell trong RAM. Một lô XML4 lớn có thể lên hàng trăm nghìn dòng.
    wb = Workbook(write_only=True)

    if summary_rows:
        # Summary đứng đầu tiên — người dùng mở file là thấy tổng quan ngay,
        # không phải lội qua các sheet dữ liệu thô trước.
        ws = wb.create_sheet(title=sanitize_sheet_name(SHEET_SUMMARY, default="Sheet"))
        summary_columns = [STUDY_ID_COLUMN] + list(SUMMARY_COLUMNS)
        append_row_as_text(ws, summary_columns)
        for rec in summary_rows:
            append_row_as_text(ws, [rec.get(c, "") for c in summary_columns])

    for title, records, lead_column, hidden_columns in sheets:
        if not records:
            continue
        ws = wb.create_sheet(title=sanitize_sheet_name(title, default="Sheet"))
        columns = _sheet_columns(
            records, mapping_columns,
            lead_column=lead_column, hidden_columns=hidden_columns,
        )
        # append_row_as_text: giá trị xét nghiệm kiểu "=<0.5" phải giữ nguyên
        # văn bản, không được biến thành công thức Excel.
        append_row_as_text(ws, columns)
        for rec in records:
            append_row_as_text(ws, [rec.get(c, "") for c in columns])

    wb.save(output_path)
    return True, (
        f"Hoàn thành! {len(included)} bản ghi Include, "
        f"{len(unclassified)} chưa có trong danh mục, "
        f"{n_excluded} bị loại (Exclude).{mapping_note}{warning}"
    )


def run_xml_to_excel(xml_files, output_path, callback, mapping_path=None):
    """Giải mã XML4 song song, đối chiếu danh mục và xuất Excel.

    ``mapping_path`` là file Excel mapping tuỳ chọn của người dùng; bỏ trống
    thì luồng xử lý y như cũ.
    """
    try:
        success, msg = _collect_and_save(xml_files, output_path, mapping_path)
    except Exception as e:
        success, msg = False, f"Lỗi: {e}"
    callback(success, msg)
